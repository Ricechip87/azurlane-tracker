import csv
import json
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

sys.dont_write_bytecode = True

import openpyxl
from lib.growth_sheet_selection import find_sheet
from lib.growth_name_matching import build_name_keys, normalize_growth_source_name, uses_unique_equipment


ROOT = Path(__file__).resolve().parents[1]
REFERENCE_DIR = ROOT / "참고용"
OUTPUT_PATH = ROOT / "src" / "data" / "growthRecommendations.json"
REPORT_DIR = ROOT / "reports" / "growth-recommendations"

KNOWN_RECENT_UNRATED_SHIPS = [
    "브리스톨(META)",
    "셰르부르",
    "아로망슈",
    "랑트레피드",
]

SOURCES = [
    {
        "key": "main",
        "label": "메인해역",
        "audience": "벽청년 이상",
        "sheet_name": "인식각성 추천표(메인해역)",
        "report": "main",
    },
    {
        "key": "operation-siren",
        "label": "대작전",
        "audience": "벽청년 이상",
        "sheet_name": "인식각성 추천표(대작전)",
        "report": "operation-siren",
        "exclusion_marker": "업데이트 안됨",
    },
    {
        "key": "newbie",
        "label": "맨땅뉴비",
        "audience": "벽뉴비 권장",
        "sheet_name": "맨땅뉴비 추천 함순이표",
        "report": "newbie",
    },
]

TIER_RE = re.compile(r"^(SS\+|SS|S\+|S-|S|A\+|A|B\+|B|C\+|C|D\+|D)\s*(급)?")
SHIP_TYPE_HEADER_RE = re.compile(r"(DD|CL|CA|CVL|CV|BB|BC|SS)")
MAIN_EXTRA_GROUP_RE = re.compile(r"(버퍼|디버퍼|힐러)")
NEWBIE_EXTRA_GROUP_RE = re.compile(r"(상시|무딱|힐러|연구함|신규|이벤트)")
MAIN_HEALER_GROUP_RE = re.compile(r"(메인 힐러|서브 힐러)")

NAME_ALIASES = {
    "하이덴 리우": "하우덴 리우",
    "렉싱턴2": "렉싱턴Ⅱ",
    "렉싱턴 2": "렉싱턴Ⅱ",
    "라피 II": "라피Ⅱ",
    "라피 Ⅱ": "라피Ⅱ",
    "다 빈치": "레오나르도 다 빈치",
    "키로프 META": "키로프(META)",
    "키로프 META (경순)": "키로프(META)",
    "U-556 META": "U-556(META)",
    "헬레나 / META": "헬레나(META)",
    "헬레나 META": "헬레나(META)",
    "I404": "이404",
    "모모 벨리아 데빌룩": "모모 베리아 데빌룩",
    "쿠니베르티": "비토리오 쿠니베르티",
    "레골로": "아틸리오 레골로",
    "아브루치": "두카 델리 아브루치",
    "가라발디": "주세페 가리발디",
    "주세페 가라발디": "주세페 가리발디",
    "돈스코이": "드미트리 돈스코이",
    "컨스텔 레이션": "컨스텔레이션",
    "타이콘 데로가": "타이콘데로가",
    "그나이 META": "그나이제나우(META)",
    "나나 (콜)": "나나 아스타 데빌룩",
    "나나 (콜라보)": "나나 아스타 데빌룩",
    "히퍼 META": "아드미랄 히퍼(META)",
    "히퍼 (전장)": "아드미랄 히퍼",
    "뉴욜리언스": "뉴올리언스",
    "꼬마 오이겐": "꼬마 프린츠 오이겐",
    "오이겐 μ": "프린츠 오이겐(μ장비)",
    "니나 (콜)": "니나 프리드",
    "니나 (콜라보)": "니나 프리드",
    "유미아 (콜라보)": "유미아 리스펠트",
    "파르제팔": "아우구스트 폰 파르제팔",
    "브론슨": "클래런스 K 브론슨",
    "파먀티 META": "파먀티 메르쿠리야(META)",
    "나히모프": "어드미럴 나히모프",
    "파트리샤 (콜라보)": "파트리샤 아벨하임",
}

NON_NAME_FRAGMENTS = [
    "Credits to:",
    "보는법",
    "최신화 날짜",
    "이전 버전",
    "관련 코멘트",
    "우선순위",
    "편성",
    "추천",
    "대체",
    "사용법",
    "필수",
    "보스전",
]


def clean(value):
    return str(value or "").replace("\r", "").strip()


def one_line(value):
    return " / ".join(line.strip() for line in clean(value).splitlines() if line.strip())


def normalize_name(value):
    return unicodedata.normalize("NFKC", clean(value)).replace("（", "(").replace("）", ")")


def build_name_attempts(value):
    pieces = split_name_candidates(value)
    attempts = []
    for size in range(min(3, len(pieces)), 1, -1):
        for start in range(len(pieces) - size + 1):
            attempts.append({
                "name": " ".join(pieces[start:start + size]),
                "indexes": tuple(range(start, start + size)),
            })
    attempts.extend({"name": piece, "indexes": (index,)} for index, piece in enumerate(pieces))
    return pieces, attempts


def split_name_candidates(value):
    text = normalize_name(value)
    if not text or len(text) > 40:
        return []
    if any(fragment in text for fragment in NON_NAME_FRAGMENTS):
        return []

    nonempty_lines = [line.strip() for line in re.split(r"[\n\r]+", text) if line.strip()]
    if len(nonempty_lines) > 3:
        return []

    pieces = []
    for line in nonempty_lines:
        slash_parts = [part.strip() for part in re.split(r"\s*/\s*", line) if part.strip()]
        if len(slash_parts) > 1 and all(len(part) <= 18 for part in slash_parts):
            pieces.extend(slash_parts)
        else:
            pieces.append(line)
    return list(dict.fromkeys(pieces))


def load_characters():
    with (ROOT / "src" / "data" / "characters.json").open(encoding="utf-8") as file:
        characters = json.load(file)

    by_name = {
        normalize_name(character["name"]): character
        for character in characters
    }
    for character in characters:
        for key in build_name_keys(character["name"]):
            by_name.setdefault(key, character)
    return by_name


def find_workbook_path():
    matches = sorted(REFERENCE_DIR.glob("*v2.1.8*배포용의 사본*.xlsx"))
    if not matches:
        matches = sorted(REFERENCE_DIR.glob("*.xlsx"))
    for path in matches:
        if "v2.1.8" in path.name:
            return path
    raise FileNotFoundError("참고용 폴더에서 v2.1.8 원본 XLSX를 찾지 못했습니다.")


def build_merged_value_map(ws):
    values = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                values[(cell.row, cell.column)] = cell.value

    for cell_range in ws.merged_cells.ranges:
        value = ws.cell(cell_range.min_row, cell_range.min_col).value
        if value in (None, ""):
            continue
        for row in range(cell_range.min_row, cell_range.max_row + 1):
            for column in range(cell_range.min_col, cell_range.max_col + 1):
                values[(row, column)] = value
    return values


def cell_value(values, row, column):
    return clean(values.get((row, column), ""))


def is_tier_cell(value):
    first_line = next((line.strip() for line in clean(value).splitlines() if line.strip()), "")
    return bool(TIER_RE.match(first_line))


def normalize_tier(value):
    first_line = next((line.strip() for line in clean(value).splitlines() if line.strip()), "")
    match = TIER_RE.match(first_line)
    return match.group(1) if match else first_line


def is_section_header(value):
    text = clean(value)
    if not text or len(text) > 30:
        return False
    if re.search(r"(구축|경순|중순|대순|항모|경항모|전함|순전|잠수)", text):
        return True
    if re.search(r"\b(DD|CL|CA|CVL|CV|BB|BC)\b", text):
        return True
    return text in {"SS", "잠수 SS"}


def is_source_group_cell(value, source_key):
    text = clean(value)
    if not text or len(text) > 30 or is_tier_cell(text):
        return False
    if is_section_header(text):
        return True
    if source_key == "main":
        return bool(MAIN_EXTRA_GROUP_RE.search(text))
    if source_key == "newbie":
        return bool(NEWBIE_EXTRA_GROUP_RE.search(text))
    return False


def build_group_rows(ws, values, source_key, excluded_from_row=None):
    group_rows = []
    last_row_exclusive = min(ws.max_row + 1, excluded_from_row or ws.max_row + 1)
    for row in range(1, last_row_exclusive):
        groups = []
        section_count = 0
        for column in range(1, ws.max_column + 1):
            # Use real cell values here so merged headers are only anchored once.
            text = clean(ws.cell(row, column).value)
            if not is_source_group_cell(text, source_key):
                continue
            if is_section_header(text):
                section_count += 1
            groups.append({"row": row, "column": column, "label": text})

        if source_key == "main":
            groups = apply_main_healer_subgroups(ws, groups, row)

        first_text = next(
            (cell_value(values, row, column) for column in range(1, ws.max_column + 1) if cell_value(values, row, column)),
            "",
        )
        if section_count >= 2 and not is_tier_cell(first_text):
            group_rows.append({"row": row, "groups": sorted(groups, key=lambda group: group["column"])})
    return sorted(group_rows, key=lambda item: item["row"])


def apply_main_healer_subgroups(ws, groups, row):
    if not any(group["label"] == "힐러" for group in groups):
        return groups
    first_ship_group = next((group for group in groups if is_section_header(group["label"])), None)
    if not first_ship_group:
        return groups

    subgroups = []
    for column in range(1, first_ship_group["column"]):
        text = clean(ws.cell(row + 1, column).value)
        if MAIN_HEALER_GROUP_RE.search(text):
            subgroups.append({"row": row, "column": column, "label": text})

    if not subgroups:
        return groups
    return sorted([*subgroups, *[group for group in groups if group["label"] != "힐러"]], key=lambda group: group["column"])


def find_column_group(row, column, group_rows):
    selected = None
    for group_row in group_rows:
        if group_row["row"] > row:
            break
        selected = group_row
    if not selected:
        return ""

    current = None
    for group in selected["groups"]:
        if group["column"] > column:
            break
        current = group
    return current["label"] if current else selected["groups"][0]["label"]


def find_tier(values, row):
    for current_row in range(row, max(0, row - 8), -1):
        tier = cell_value(values, current_row, 1)
        if is_tier_cell(tier):
            return normalize_tier(tier)
    return ""


def is_note_like_character_name(note, by_name):
    text = clean(note)
    if not text:
        return False
    return any(key in by_name for key in build_name_keys(text))


def find_role_note(values, row, column, by_name):
    for offset in (1, 2):
        note = cell_value(values, row + offset, column)
        if not note:
            continue
        if is_tier_cell(note):
            continue
        if is_note_like_character_name(note, by_name):
            continue
        return note
    return ""


def find_source_updated_at(ws):
    pattern = re.compile(r"최신화\s*날짜\s*:\s*(\d{4})[./-](\d{1,2})[./-](\d{1,2})")
    for row in ws.iter_rows():
        for cell in row:
            match = pattern.search(clean(cell.value))
            if match:
                year, month, day = (int(part) for part in match.groups())
                return f"{year:04d}-{month:02d}-{day:02d}"
    return ""


def find_excluded_from_row(ws, marker):
    if not marker:
        return None
    for row in ws.iter_rows():
        if any(clean(cell.value) == marker for cell in row):
            return row[0].row
    return None


def match_character(name, by_name):
    source_name = normalize_growth_source_name(name)
    lookup = NAME_ALIASES.get(source_name, source_name)
    if lookup.endswith(" μ"):
        lookup = f"{lookup[:-2]}(μ장비)"
    for key in build_name_keys(lookup):
        if key in by_name:
            return by_name[key]
    return None


def extract_source(wb, source, by_name):
    ws = find_sheet(wb, source["sheet_name"])
    values = build_merged_value_map(ws)
    exclusion_marker = source.get("exclusion_marker")
    excluded_from_row = find_excluded_from_row(ws, exclusion_marker)
    group_rows = build_group_rows(ws, values, source["key"], excluded_from_row)
    candidates = []
    unmatched = []

    last_row_exclusive = min(ws.max_row + 1, excluded_from_row or ws.max_row + 1)
    for row in range(1, last_row_exclusive):
        for column in range(1, ws.max_column + 1):
            raw = clean(ws.cell(row, column).value)
            if not raw:
                continue
            pieces, attempts = build_name_attempts(raw)
            consumed_indexes = set()
            matched_ids = set()
            for attempt in attempts:
                if any(index in consumed_indexes for index in attempt["indexes"]):
                    continue
                name = attempt["name"]
                character = match_character(name, by_name)
                if not character:
                    continue

                character_id = str(character["id"])
                if character_id in matched_ids:
                    continue
                matched_ids.add(character_id)
                consumed_indexes.update(attempt["indexes"])

                candidates.append({
                    "source": source["key"],
                    "sourceLabel": source["label"],
                    "audience": source["audience"],
                    "id": character["id"],
                    "name": character["name"],
                    "rarity": character["rarity"],
                    "faction": character["faction"],
                    "shipType": character["shipType"],
                    "tier": find_tier(values, row),
                    "sheetGroup": find_column_group(row, column, group_rows),
                    "roleNote": find_role_note(values, row, column, by_name),
                    "requiresUniqueEquipment": uses_unique_equipment(name),
                    "row": row,
                    "column": column,
                })

            for index, name in enumerate(pieces):
                if index in consumed_indexes or match_character(name, by_name):
                    continue
                if 1 < len(name) <= 20 and re.search(r"[가-힣A-Za-z0-9]", name):
                    unmatched.append({
                        "source": source["key"],
                        "name": name,
                        "row": row,
                        "column": column,
                    })

    matched_ids_by_row = defaultdict(set)
    for candidate in candidates:
        matched_ids_by_row[candidate["row"]].add(str(candidate["id"]))

    recommendations = []
    for candidate in candidates:
        if len(matched_ids_by_row[candidate["row"]]) >= 2:
            recommendations.append(candidate)
        else:
            unmatched.append({
                "source": source["key"],
                "name": candidate["name"],
                "row": candidate["row"],
                "column": candidate["column"],
                "reason": "single matched name row",
            })

    return {
        "source": source,
        "sheet": ws.title.strip(),
        "updatedAt": find_source_updated_at(ws),
        "exclusionMarker": exclusion_marker,
        "excludedFromRow": excluded_from_row,
        "rows": ws.max_row,
        "columns": ws.max_column,
        "recommendations": recommendations,
        "unmatched": unmatched,
    }


def dedupe_recommendations(recommendations):
    seen = set()
    output = []
    for item in recommendations:
        key = "|".join(str(item.get(part, "")) for part in [
            "source",
            "id",
            "tier",
            "sheetGroup",
            "roleNote",
            "requiresUniqueEquipment",
            "row",
            "column",
        ])
        if key in seen:
            continue
        seen.add(key)
        output.append(item)
    return output


def write_csv(path, rows):
    with path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.writer(file)
        for row in rows:
            writer.writerow([
                "\n".join(line.rstrip() for line in value.splitlines()) if isinstance(value, str) else value
                for value in row
            ])


def group_by(items, key):
    grouped = defaultdict(list)
    for item in items:
        grouped[key(item)].append(item)
    return grouped


def write_recommendation_text_report(path, items):
    lines = [
        "육성 추천표 자동 추출 결과",
        "",
        f"총 추천 엔트리: {len(items)}",
        "표기: 함선명 [등급/진영/함종] - 역할 메모 (원본 row:column)",
        "",
    ]
    for source_label, source_items in group_by(items, lambda item: f"{item['sourceLabel']} / {item['audience']}").items():
        lines.extend([f"## {source_label}", ""])
        for tier, tier_items in group_by(source_items, lambda item: item.get("tier") or "티어 미확인").items():
            lines.append(f"### {tier}")
            for sheet_group, group_items in group_by(tier_items, lambda item: item.get("sheetGroup") or "분류 미확인").items():
                lines.append(f"- {sheet_group}")
                for item in group_items:
                    note = one_line(item.get("roleNote", ""))
                    if item.get("requiresUniqueEquipment"):
                        prefix = "\uc804\uc6a9\uc7a5\ube44 \uae30\uc900"
                        note = f"{prefix} / {note}" if note else prefix
                    note_text = f" - {note}" if note else ""
                    lines.append(
                        f"  - {item['name']} [{item['rarity']}/{item['faction']}/{item['shipType']}]"
                        f"{note_text} ({item['row']}:{item['column']})"
                    )
            lines.append("")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_unmatched_text_report(path, items):
    lines = [
        "육성 추천표 미매칭 검토 후보",
        "",
        f"총 검토 후보: {len(items)}",
        "",
    ]
    for source, source_items in group_by(items, lambda item: item["source"]).items():
        lines.append(f"## {source}")
        for item in source_items:
            lines.append(f"- {item['name']} ({item['row']}:{item['column']}) - {item.get('reason', 'not matched')}")
        lines.append("")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_audit_report(path, extracted, recommendations, unmatched, unrated_recent_ships):
    lines = [
        "# 육성 추천 데이터 감사",
        "",
        "## 원본 범위",
        "",
        "| 구분 | 원본 최신화 | 추천 엔트리 | 고유 함선 |",
        "|---|---:|---:|---:|",
    ]
    for result in extracted:
        source_key = result["source"]["key"]
        items = [item for item in recommendations if item["source"] == source_key]
        lines.append(
            f"| {result['source']['label']} | {result['updatedAt'] or '미확인'} | "
            f"{len(items)} | {len({item['name'] for item in items})} |"
        )

    blank_tiers = [item for item in recommendations if not item.get("tier")]
    blank_groups = [item for item in recommendations if not item.get("sheetGroup")]
    blank_notes = [item for item in recommendations if not item.get("roleNote")]
    recommendation_rows = {(item["source"], item["row"]) for item in recommendations}
    focused_unmatched = [
        item for item in unmatched
        if (item["source"], item["row"]) in recommendation_rows and item.get("reason", "not matched") == "not matched"
    ]
    operation_source = next(result for result in extracted if result["source"]["key"] == "operation-siren")
    exclusion_marker = operation_source.get("exclusionMarker") or "미확인"
    excluded_from_row = operation_source.get("excludedFromRow")
    lines.extend([
        "",
        "## 구조 검사",
        "",
        f"- 빈 티어: {len(blank_tiers)}건",
        f"- 빈 분류: {len(blank_groups)}건",
        f"- 빈 역할 메모: {len(blank_notes)}건",
        f"- 실제 추천 행에 남은 미매칭 문구: {len(focused_unmatched)}건",
        f"- 대작전 `{exclusion_marker}` 표식 구간: 추출 제외"
        f" (현재 {excluded_from_row}행부터)" if excluded_from_row else "- 대작전 제외 표식: 찾지 못함",
        "",
        f"대작전 원본에서 `{exclusion_marker}` 표식이 시작되는 구간은 "
        "사용자 검토에 따라 앱 JSON과 추천 순위에서 제외합니다. 행 번호는 원본 구조에 따라 자동으로 다시 찾습니다.",
        "",
        "## 원본 미평가 신규 함선",
        "",
    ])
    if unrated_recent_ships:
        lines.extend(f"- {item['name']} ({item['id']})" for item in unrated_recent_ships)
    else:
        lines.append("- 없음")

    if blank_notes:
        lines.extend(["", "## 역할 메모가 빈 원본 엔트리", ""])
        lines.extend(
            f"- {item['source']} / {item['name']} ({item['row']}:{item['column']})"
            for item in blank_notes
        )

    if focused_unmatched:
        lines.extend(["", "## 추천 행에 남은 미매칭 문구", ""])
        lines.extend(
            f"- {item['source']} / {item['name']} ({item['row']}:{item['column']})"
            for item in focused_unmatched
        )

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main():
    workbook_path = find_workbook_path()
    by_name = load_characters()
    wb = openpyxl.load_workbook(workbook_path, read_only=False, data_only=True)

    extracted = [extract_source(wb, source, by_name) for source in SOURCES]
    recommendations = dedupe_recommendations([item for result in extracted for item in result["recommendations"]])
    unmatched = [item for result in extracted for item in result["unmatched"]]
    recommendation_names = {item["name"] for item in recommendations}
    unrated_recent_ships = []
    for name in KNOWN_RECENT_UNRATED_SHIPS:
        character = match_character(name, by_name)
        if character and character["name"] not in recommendation_names:
            unrated_recent_ships.append({
                "id": character["id"],
                "name": character["name"],
                "reason": "원본 최신화 이후 미평가",
            })

    output = {
        "notes": [
            "Generated from the original XLSX workbook in 참고용.",
            "Reference files are read-only inputs; edit the extraction script or source workbook to regenerate.",
            "row and column are 1-based XLSX coordinates for manual review.",
        ],
        "sources": [
            {
                "key": source["key"],
                "label": source["label"],
                "audience": source["audience"],
                "file": workbook_path.name,
                "sheet": extracted[index]["sheet"],
                "updatedAt": extracted[index]["updatedAt"],
                "exclusionMarker": extracted[index]["exclusionMarker"],
                "excludedFromRow": extracted[index]["excludedFromRow"],
            }
            for index, source in enumerate(SOURCES)
        ],
        "recommendations": recommendations,
        "review": {
            "unmatched": unmatched[:300],
            "unmatchedTotal": len(unmatched),
            "unratedRecentShips": unrated_recent_ships,
        },
    }

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(output, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    write_csv(
        REPORT_DIR / "review.csv",
        [
            ["source", "sourceLabel", "audience", "tier", "sheetGroup", "name", "id", "rarity", "faction", "shipType", "roleNote", "requiresUniqueEquipment", "row", "column"],
            *[
                [
                    item["source"],
                    item["sourceLabel"],
                    item["audience"],
                    item["tier"],
                    item["sheetGroup"],
                    item["name"],
                    item["id"],
                    item["rarity"],
                    item["faction"],
                    item["shipType"],
                    item["roleNote"],
                    item["requiresUniqueEquipment"],
                    item["row"],
                    item["column"],
                ]
                for item in recommendations
            ],
        ],
    )
    write_csv(
        REPORT_DIR / "unmatched.csv",
        [
            ["source", "name", "row", "column", "reason"],
            *[
                [
                    item["source"],
                    item["name"],
                    item["row"],
                    item["column"],
                    item.get("reason", "not matched"),
                ]
                for item in unmatched
            ],
        ],
    )
    write_recommendation_text_report(REPORT_DIR / "review.txt", recommendations)
    write_unmatched_text_report(REPORT_DIR / "unmatched.txt", unmatched)
    write_audit_report(REPORT_DIR / "audit.md", extracted, recommendations, unmatched, unrated_recent_ships)

    for source in SOURCES:
        report_name = source["report"]
        write_recommendation_text_report(
            REPORT_DIR / f"{report_name}.review.txt",
            [item for item in recommendations if item["source"] == source["key"]],
        )
        write_unmatched_text_report(
            REPORT_DIR / f"{report_name}.unmatched.txt",
            [item for item in unmatched if item["source"] == source["key"]],
        )

    for result in extracted:
        print(
            f"{result['source']['label']}: rows={result['rows']} columns={result['columns']} "
            f"matched={len(result['recommendations'])} unmatchedCandidates={len(result['unmatched'])}"
        )
    print(f"totalRecommendations={len(recommendations)}")
    print(f"reviewUnmatchedTotal={len(unmatched)}")
    print(f"wrote {OUTPUT_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
